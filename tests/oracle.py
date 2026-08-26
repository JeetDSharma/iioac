"""Reference implementation of the fugitive chain, driven by workbook.py.

Reads unit values from the ORIGINAL AERMOD .xlsx inside IIOAC_RunFiles.zip via
openpyxl, so a run through this oracle exercises a completely separate data path
from the .bin files the web app consumes.
"""
import io
import os
import zipfile

import openpyxl

import workbook as W

HERE = os.path.dirname(os.path.abspath(__file__))

# AERMOD run files, fetched and checksum-verified by fetch_fixtures.py. The default
# fixture holds only the 18 files the fidelity and sampling suites need (~10 MB).
# Set IIOAC_RUNFILES to EPA's full IIOAC_RunFiles.zip to run the 672-file sweep in
# test_data.py. See tests/README.md.
RUNFILES = os.environ.get('IIOAC_RUNFILES') or os.path.join(HERE, 'fixtures', 'runfiles.zip')
if not os.path.exists(RUNFILES):
    raise SystemExit(
        f'run files not found: {RUNFILES}\n'
        'Run tests/fetch_fixtures.py, or set IIOAC_RUNFILES to a copy of\n'
        'IIOAC_RunFiles.zip. See tests/README.md.')

BAND_ORDER = ['inner', 'outer', 'all', 'community']


def load_xlsx_grid(name):
    """1460 x 60 floats from an AERMOD results workbook, read with openpyxl."""
    with zipfile.ZipFile(RUNFILES) as z:
        data = z.read(name)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=1461, max_col=60, values_only=True)):
        grid.append([float(v) if v not in (None, '') else 0.0 for v in row])
    assert len(grid) == 1460, f'{name}: got {len(grid)} rows'
    return grid


def unit(grid, band, frequency, col):
    return grid[BAND_ORDER.index(band) * 365 + frequency - 1][col]


def scaling_factors(area, particle_label, locale):
    coeffs = W.fugitive_scaling_coefficients()[particle_label + locale]
    ref = W.scaling_reference_area()
    out = {}
    for band, ab in coeffs.items():
        a, b = ab['a'], ab['b']
        out[band] = (a * area ** b) / (a * ref ** b)
    return out


def calculate(grid, scenarios, area, particle_label, locale):
    """Returns {(stat, location): {...}} mirroring 'Fugitive Output Prelim' rows 5-10."""
    row_map = W.calc_row_map()
    labels = W.calc_row_labels()
    out_rows = W.output_row_map()
    rate_const = W.emission_rate_constant()
    ratios = W.io_ratios()
    caps = W.fugitive_caps()[particle_label]
    sf = scaling_factors(area, particle_label, locale)
    receptors = W.receptor_table()
    coeffs = W.dose_formula_coefficients()

    # Which cap governs which calculation row, following LoopCalcs' block ranges.
    def cap_for(calc_row):
        if 2 <= calc_row <= 13:
            return caps['dailyAir']
        if 14 <= calc_row <= 25:
            return caps['annualAir']
        return None   # 26-61: deposition blocks, all uncapped in the workbook

    def calc_cell(calc_row):
        col_base, band = row_map[calc_row]
        cap = cap_for(calc_row)
        total = 0.0
        for s in scenarios:
            rate = s['amount'] / s['hours'] * rate_const
            v = rate * unit(grid, band, s['frequency'], col_base + s['durationCol'])
            if cap is not None:
                v *= sf[band]
                if v > cap:
                    v = cap
            total += v
        return total

    results = {}
    for r, spec in out_rows.items():
        od = calc_cell(spec['outdoorDaily'])
        oa = calc_cell(spec['outdoorAnnual'])
        ratio = ratios[spec['ioRatioCell']]
        idl, ial = od * ratio, oa * ratio

        acute_mul, acute_div = coeffs['acute']
        chronic_mul, chronic_div = coeffs['chronic']
        life_mul, life_div = coeffs['lifetime']

        def prod(xs):
            p = 1.0
            for x in xs:
                p *= x
            return p

        acute, chronic = [], []
        for rec in receptors[:7]:
            acute.append((rec['pctIndoor'] * idl + rec['pctOutdoor'] * od)
                         * rec['inhAcute'] * prod(acute_mul) / (rec['bw'] * prod(acute_div)))
            chronic.append((rec['pctIndoor'] * ial + rec['pctOutdoor'] * oa)
                           * rec['inhChronic'] * prod(chronic_mul) / (rec['bw'] * prod(chronic_div)))
        life = receptors[7]
        lifetime = ((life['pctIndoor'] * ial + life['pctOutdoor'] * oa)
                    * life['inhChronic'] * prod(life_mul) / (life['bw'] * prod(life_div)))

        results[(spec['stat'], spec['location'])] = {
            'outdoorDaily': od, 'outdoorAnnual': oa,
            'indoorDaily': idl, 'indoorAnnual': ial,
            'totalDep': calc_cell(spec['totalDep']),
            'wetDep': calc_cell(spec['wetDep']),
            'dryDep': calc_cell(spec['dryDep']),
            'acute': acute, 'chronic': chronic, 'lifetime': lifetime,
        }
    return results, sf, labels
