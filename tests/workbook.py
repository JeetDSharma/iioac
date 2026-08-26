"""Read the model's structure straight out of iioac_1.0.xlsm.

Nothing here is transcribed by hand. The row -> (column group, receptor band) map
and the dose coefficients are parsed from the workbook's own formula strings, so
this module is an independent oracle for the JavaScript engine rather than a
restatement of it.
"""
import os
import re

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))

# EPA's workbook, fetched and checksum-verified by fetch_fixtures.py. Point
# IIOAC_WORKBOOK at your own copy to verify against a download you made yourself.
WORKBOOK = os.environ.get('IIOAC_WORKBOOK') or os.path.join(HERE, 'fixtures', 'iioac_1.0.xlsm')
if not os.path.exists(WORKBOOK):
    raise SystemExit(
        f'workbook not found: {WORKBOOK}\n'
        'Run tests/fetch_fixtures.py, or set IIOAC_WORKBOOK to a copy of\n'
        'IIOAC 1.0.xlsm. See tests/README.md.')

# Row where each receptor band starts in 'Point+Fugitive Lookup'.
BAND_START = {4: 'inner', 369: 'outer', 734: 'all', 1099: 'community'}


def col_index(letters):
    n = 0
    for ch in letters:
        n = n * 26 + ord(ch) - 64
    return n


_wb = None


def wb():
    global _wb
    if _wb is None:
        _wb = openpyxl.load_workbook(WORKBOOK, data_only=False)
    return _wb


def calc_row_map(sheet='Fugitive Calculations Prelim'):
    """rows 2..61 -> (column offset into the 60-column run file, band name).

    Parsed from the INDEX(...) range in each row's column-E formula.
    """
    ws = wb()[sheet]
    out = {}
    for r in range(2, 62):
        f = ws.cell(r, 5).value
        if not f:
            continue
        m = re.search(r"Lookup'!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)", f)
        if not m:
            raise AssertionError(f'{sheet} row {r}: no INDEX range in formula')
        # Column D of the lookup sheet is the first of the 60 data columns.
        out[r] = (col_index(m.group(1)) - 4, BAND_START[int(m.group(2))])
    return out


def calc_row_labels(sheet='Fugitive Calculations Prelim'):
    """rows 2..61 -> (location label, statistic label) from columns C and D."""
    ws = wb()[sheet]
    return {r: (ws.cell(r, 3).value, ws.cell(r, 4).value) for r in range(2, 62)}


def emission_rate_constant(sheet='Fugitive Calculations Prelim'):
    """The kg/day-over-N-hours -> g/s constant, read from the formula text."""
    f = wb()[sheet]['E2'].value
    m = re.search(r'\*\s*([0-9.]+)\s*\)', f)
    if not m:
        raise AssertionError('could not find the emission rate constant')
    return float(m.group(1))


def output_row_map(sheet='Fugitive Output Prelim'):
    """Output rows 5..10 -> {parameter: calculation-sheet row}."""
    ws = wb()[sheet]
    cols = {'outdoorDaily': 3, 'outdoorAnnual': 4, 'totalDep': 7, 'wetDep': 8, 'dryDep': 9}
    out = {}
    stat = None
    for r in range(5, 11):
        # Column A is merged across each statistic's three location rows.
        stat = ws.cell(r, 1).value or stat
        entry = {'stat': stat, 'location': ws.cell(r, 2).value}
        for name, c in cols.items():
            f = ws.cell(r, c).value
            m = re.search(r"Calculations'!\$E\$?(\d+):\$?[A-Z]+\$?(\d+)", str(f))
            if not m:
                raise AssertionError(f'{sheet} {ws.cell(r, c).coordinate}: unparsed {f!r}')
            assert m.group(1) == m.group(2), f'row mismatch in {f!r}'
            entry[name] = int(m.group(1))
        # Indoor = outdoor * lookups!B57 (mean) or B58 (high end).
        m = re.search(r'lookups!\$B\$(\d+)', str(ws.cell(r, 5).value))
        entry['ioRatioCell'] = f'B{m.group(1)}'
        out[r] = entry
    return out


def io_ratios():
    ws = wb()['lookups']
    return {'B57': ws['B57'].value, 'B58': ws['B58'].value}


def _eval_lookup(ref):
    """Resolve a lookups!<cell> reference, evaluating the sheet's own formulas."""
    ws = wb()['lookups']
    v = ws[ref].value
    if isinstance(v, str) and v.startswith('='):
        expr = v[1:]
        expr = re.sub(r'\bB(\d+)\b', lambda m: f'({_eval_lookup("B" + m.group(1))})', expr)
        expr = re.sub(r'\bC(\d+)\b', lambda m: f'({_eval_lookup("C" + m.group(1))})', expr)
        expr = re.sub(r'\bD(\d+)\b', lambda m: f'({_eval_lookup("D" + m.group(1))})', expr)
        return eval(expr, {'__builtins__': {}}, {})
    return v


def receptor_table():
    """lookups!A25:F32 with the formula cells evaluated."""
    ws = wb()['lookups']
    rows = []
    for r in range(25, 33):
        rows.append({
            'name': ws.cell(r, 1).value,
            'bw': _eval_lookup(f'B{r}'),
            'inhAcute': _eval_lookup(f'C{r}'),
            'inhChronic': _eval_lookup(f'D{r}'),
            'pctIndoor': ws.cell(r, 5).value,
            'pctOutdoor': ws.cell(r, 6).value,
        })
    return rows


def dose_formula_coefficients(sheet='Fugitive Output Prelim'):
    """Pull the literal multipliers and divisors out of the dose formulas.

    Acute:   ((%in*indoorDaily + %out*outdoorDaily) * InhRacute * 24 * 1 * 0.001) / (BW * 1)
    Chronic: ((%in*indoorAnn   + %out*outdoorAnn)   * InhRchronic  * 1 * 0.001) / (BW * 1)
    Lifetime uses 33 and divides by BW * 78.
    """
    ws = wb()[sheet]
    def parse(cell):
        f = ws[cell].value
        nums = re.findall(r'\*\s*([0-9.]+)\s*', f.split(')/(')[0])
        denom = re.findall(r'\*\s*([0-9.]+)\s*\)', f.split(')/(')[1])
        return [float(x) for x in nums], [float(x) for x in denom]
    return {'acute': parse('J5'), 'chronic': parse('Q5'), 'lifetime': parse('X5')}


def fugitive_caps():
    """lookups!B60 block: air concentration caps by particle size, fugitive column."""
    ws = wb()['lookups']
    return {
        'Coarse': {'dailyAir': ws['C63'].value, 'annualAir': ws['C64'].value},
        'Fine': {'dailyAir': ws['E63'].value, 'annualAir': ws['E64'].value},
        'No particles (vapor only)': {'dailyAir': ws['G63'].value, 'annualAir': ws['G64'].value},
    }


def fugitive_scaling_coefficients():
    """lookups!A43:E54 -> {particle+locale: {band: (a, b)}}."""
    ws = wb()['lookups']
    cols = {'inner': 2, 'outer': 3, 'community': 4, 'all': 5}
    out = {}
    for r in range(43, 55):
        key = ws.cell(r, 1).value
        which = key[-1]
        base = key[:-1]
        entry = out.setdefault(base, {})
        for band, c in cols.items():
            entry.setdefault(band, {})[which] = ws.cell(r, c).value
    return out


def scaling_reference_area(sheet='lookups', cell='B72'):
    """The denominator area in the fugitive scaling formula (100 m2)."""
    f = wb()[sheet][cell].value
    m = re.search(r'\*\s*(\d+)\s*\^VLOOKUP', f)
    return float(m.group(1))
